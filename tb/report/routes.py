"""Dashboard and adherence report routes."""
from __future__ import annotations

import csv
from calendar import monthrange
from datetime import date
from io import BytesIO, StringIO

import openpyxl
from openpyxl.styles import Font, PatternFill
from flask import Blueprint, render_template, request, send_file
from sqlalchemy import case as sa_case
from sqlalchemy import func

from tb.adherence import (
    get_adherence_stats_bulk,
    get_at_risk_patients,
    get_monthly_adherence_trend,
    get_tb_type_stats,
)
from tb.constants import OUTCOME_LABELS
from tb.extensions import db
from tb.models import MedicationDose, Patient
from tb.security import staff_required
from tb.time_utils import today_th

bp = Blueprint("report", __name__)


@bp.route("/dashboard")
@staff_required
def dashboard():
    patients = Patient.query.filter_by(archived=False).order_by(Patient.id).all()
    today = today_th()
    patient_ids = [p.id for p in patients]
    stats_map = get_adherence_stats_bulk(patient_ids, today)
    today_doses = {}
    if patient_ids:
        today_doses = {
            d.patient_id: d for d in MedicationDose.query.filter(
                MedicationDose.patient_id.in_(patient_ids),
                MedicationDose.date == today,
            ).all()
        }

    rows = []
    total_overdue_all = 0
    missed_today = []
    for p in patients:
        stats = stats_map[p.id]
        today_dose = today_doses.get(p.id)
        rows.append({"patient": p, "stats": stats, "today_dose": today_dose})
        total_overdue_all += stats["overdue"]
        if today_dose and not today_dose.taken:
            missed_today.append(p)

    at_risk = get_at_risk_patients(patients, stats_map, today)
    tb_type_stats = get_tb_type_stats(patients, stats_map)

    return render_template(
        "dashboard.html", rows=rows, today=today,
        total_patients=len(patients), total_overdue=total_overdue_all,
        missed_today=missed_today,
        scheduled_today_count=len(today_doses),
        at_risk=at_risk,
        tb_type_stats=tb_type_stats,
    )


@bp.route("/report")
@staff_required
def report():
    today = today_th()
    year = request.args.get("year", today.year, type=int)
    month = request.args.get("month", today.month, type=int)
    _, last_day = monthrange(year, month)
    first = date(year, month, 1)
    last = date(year, month, last_day)

    patients = Patient.query.filter_by(archived=False).order_by(Patient.id).all()
    patient_ids = [p.id for p in patients]

    rows_q = db.session.query(
        MedicationDose.patient_id,
        func.count(MedicationDose.id).label("total"),
        func.sum(sa_case((MedicationDose.taken == True, 1), else_=0)).label("taken"),
    ).filter(
        MedicationDose.patient_id.in_(patient_ids),
        MedicationDose.date >= first,
        MedicationDose.date <= last,
    ).group_by(MedicationDose.patient_id).all()

    stats_by_pid = {r.patient_id: r for r in rows_q}
    report_rows = []
    for p in patients:
        r = stats_by_pid.get(p.id)
        total = r.total if r else 0
        taken = r.taken if r else 0
        overdue = total - taken
        pct = round(taken / total * 100, 1) if total else 0
        report_rows.append({
            "patient": p,
            "total": total,
            "taken": taken,
            "overdue": overdue,
            "pct": pct,
        })

    avg_pct = round(
        sum(row["pct"] for row in report_rows if row["total"] > 0) /
        max(sum(1 for row in report_rows if row["total"] > 0), 1),
        1,
    )

    month_options = []
    for delta in range(-11, 4):
        y = today.year + (today.month - 1 + delta) // 12
        m = (today.month - 1 + delta) % 12 + 1
        month_options.append((y, m))

    trend_data = get_monthly_adherence_trend(patient_ids, num_months=6)

    return render_template(
        "report.html",
        report_rows=report_rows, year=year, month=month,
        month_options=month_options, avg_pct=avg_pct,
        total_patients=len([r for r in report_rows if r["total"] > 0]),
        trend_data=trend_data,
    )


@bp.route("/report/export")
@staff_required
def report_export():
    today = today_th()
    year = request.args.get("year", today.year, type=int)
    month = request.args.get("month", today.month, type=int)
    _, last_day = monthrange(year, month)
    first = date(year, month, 1)
    last = date(year, month, last_day)

    patients = Patient.query.filter_by(archived=False).order_by(Patient.id).all()
    patient_ids = [p.id for p in patients]

    rows_q = db.session.query(
        MedicationDose.patient_id,
        func.count(MedicationDose.id).label("total"),
        func.sum(sa_case((MedicationDose.taken == True, 1), else_=0)).label("taken"),
    ).filter(
        MedicationDose.patient_id.in_(patient_ids),
        MedicationDose.date >= first,
        MedicationDose.date <= last,
    ).group_by(MedicationDose.patient_id).all()

    stats_by_pid = {r.patient_id: r for r in rows_q}
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ชื่อ", "HN", "ชนิด TB", "โดสในเดือน", "กินแล้ว", "ค้าง",
        "% ความสม่ำเสมอ", "ผลการรักษา",
    ])
    for p in patients:
        r = stats_by_pid.get(p.id)
        total = r.total if r else 0
        taken = r.taken if r else 0
        pct = round(taken / total * 100, 1) if total else 0
        outcome_label = OUTCOME_LABELS.get(p.outcome or "", "กำลังรักษา")
        writer.writerow([
            p.name, p.hn, p.tb_type, total, taken, total - taken, pct, outcome_label,
        ])
    output.seek(0)
    filename = f"adherence_report_{year}_{month:02d}.csv"
    return send_file(
        BytesIO(output.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename,
    )


@bp.route("/report/export/xlsx")
@staff_required
def report_export_xlsx():
    today = today_th()
    year = request.args.get("year", today.year, type=int)
    month = request.args.get("month", today.month, type=int)
    _, last_day = monthrange(year, month)
    first = date(year, month, 1)
    last = date(year, month, last_day)

    patients = Patient.query.filter_by(archived=False).order_by(Patient.id).all()
    patient_ids = [p.id for p in patients]

    rows_q = db.session.query(
        MedicationDose.patient_id,
        func.count(MedicationDose.id).label("total"),
        func.sum(sa_case((MedicationDose.taken == True, 1), else_=0)).label("taken"),
    ).filter(
        MedicationDose.patient_id.in_(patient_ids),
        MedicationDose.date >= first,
        MedicationDose.date <= last,
    ).group_by(MedicationDose.patient_id).all()

    stats_by_pid = {r.patient_id: r for r in rows_q}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    headers = ["ชื่อ", "HN", "TB No.", "ชนิด TB", "อายุ", "น้ำหนัก (kg)",
               "วันเริ่มรักษา", "โดสในเดือน", "กินแล้ว", "ค้าง", "% Adherence", "ผลการรักษา"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    fill_green = PatternFill("solid", fgColor="C8E6C9")
    fill_yellow = PatternFill("solid", fgColor="FFF9C4")
    fill_red = PatternFill("solid", fgColor="FFCDD2")

    for p in patients:
        r = stats_by_pid.get(p.id)
        total = r.total if r else 0
        taken = r.taken if r else 0
        overdue = total - taken
        pct = round(taken / total * 100, 1) if total else 0
        outcome_label = OUTCOME_LABELS.get(p.outcome or "", "กำลังรักษา")
        ws.append([
            p.name, p.hn, p.tb_no, p.tb_type, p.age, p.weight,
            p.start_date.strftime("%Y-%m-%d"), total, taken, overdue, pct, outcome_label,
        ])
        pct_cell = ws.cell(row=ws.max_row, column=11)
        if total > 0:
            pct_cell.fill = fill_green if pct >= 90 else (fill_yellow if pct >= 70 else fill_red)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"tb_report_{year}_{month:02d}.xlsx",
    )


@bp.route("/ping")
def ping():
    return "ok", 200
